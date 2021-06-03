#coding:utf-8
import openpyxl
from datetime import datetime
from openpyxl.styles import Border,Side

#基础参数设置
file_name="AAA.xlsx"#文件名字
Every_next=3	#几个考场换行
kao_name="高二年级5月考试考场安排"

print("开始筛选数据")
#设置班级或者考场的最大人数，最好Max+1
max_class=62
max_kc=43

#数组下标从0开始
wb = openpyxl.load_workbook(file_name)
#ws=wb['Sheet1']
ws = wb.active

sht_bj= wb.create_sheet("按班级")
sht_kc= wb.create_sheet("按考场")

rows=ws.max_row   #获取行数
cols=ws.max_column    #获取列数

#班级人数
Class_num=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,]
#考场人数
KaoC_num=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,]
#0用于标记考场最大值
a=2
#遍历取得最大班级和考场数量、考场和班级人数最大值
while a<rows:
    bj= int(ws.cell(a,2).value)
    kc= int(ws.cell(a,3).value)
    #找到班级，将一行数据复制走，继续搜索
    if(Class_num[0]<bj):
        Class_num[0]=bj
    if(KaoC_num[0]<kc):
        KaoC_num[0]=kc
    a=a+1
print(Class_num[0])
print(KaoC_num[0])

#设置边框
border = Border(left=Side(border_style='thin',color='000000'),
			right=Side(border_style='thin',color='000000'),
			top=Side(border_style='thin',color='000000'),
			bottom=Side(border_style='thin',color='000000'))

#搜索所有记录，将对应班级和考场复制到相应区域
i=2#标记行号，从第二行开始计算
while i<=rows:
	#第二列是班级
	bj_value= ws.cell(i,2).value
	bj_value=int(bj_value)
	Class_num[bj_value]=Class_num[bj_value]+1
	
	#复制到对应工作簿的对应位置
	bj_s=bj_value//Every_next
	bj_ys=bj_value%Every_next
	if(bj_ys==0):
		bj_s=bj_s-1
		bj_ys=Every_next
	row=int(bj_s*max_class+Class_num[bj_value])
	#print(bj_ys*5)
	sht_bj.cell(row+1,bj_ys*5-4).value=ws.cell(i,1).value
	sht_bj.cell(row+1,bj_ys*5-3).value=ws.cell(i,2).value
	sht_bj.cell(row+1,bj_ys*5-2).value=ws.cell(i,3).value#继续处理考场
	sht_bj.cell(row+1,bj_ys*5-1).value=ws.cell(i,4).value
	#添加边框
	sht_bj.cell(row+1,bj_ys*5-4).border=border
	sht_bj.cell(row+1,bj_ys*5-3).border=border
	sht_bj.cell(row+1,bj_ys*5-2).border=border
	sht_bj.cell(row+1,bj_ys*5-1).border=border
	
	kc_value= ws.cell(i,3).value
	kc_value=int(kc_value)
	KaoC_num[kc_value]=KaoC_num[kc_value]+1
	kc_s=kc_value//Every_next
	kc_ys=kc_value%Every_next
	if(kc_ys==0):
		kc_s=kc_s-1
		kc_ys=Every_next
	row=int(kc_s*max_kc+KaoC_num[kc_value])
	#复制过去
	sht_kc.cell(row+1,kc_ys*5-4).value=ws.cell(i,1).value
	sht_kc.cell(row+1,kc_ys*5-3).value=ws.cell(i,2).value
	sht_kc.cell(row+1,kc_ys*5-2).value=ws.cell(i,3).value
	sht_kc.cell(row+1,kc_ys*5-1).value=ws.cell(i,4).value
	#添加边框
	sht_kc.cell(row+1,kc_ys*5-4).border=border
	sht_kc.cell(row+1,kc_ys*5-3).border=border
	sht_kc.cell(row+1,kc_ys*5-2).border=border
	sht_kc.cell(row+1,kc_ys*5-1).border=border

	kc_value=int(kc_value)
	#找到班级，将一行数据复制走，继续搜索
	i=i+1

i=1#标记行号，从第二行开始计算
sum=0
while i<=KaoC_num[0]:
	#第i考场数据处理
    kc_value=i
    sum=sum+KaoC_num[i]
    kc_s=kc_value//Every_next
    kc_ys=kc_value%Every_next
    if(kc_ys==0):
        kc_s=kc_s-1
        kc_ys=Every_next
    row=int(kc_s*max_kc)
    sht_kc.cell(row+1,kc_ys*5-4).value="姓名"
    sht_kc.cell(row+1,kc_ys*5-3).value="班级"
    sht_kc.cell(row+1,kc_ys*5-2).value="考场"
    sht_kc.cell(row+1,kc_ys*5-1).value="考号"
	#设置边框
    sht_kc.cell(row+1,kc_ys*5-4).border=border
    sht_kc.cell(row+1,kc_ys*5-3).border=border
    sht_kc.cell(row+1,kc_ys*5-2).border=border
    sht_kc.cell(row+1,kc_ys*5-1).border=border
    #print(KaoC_num[i])
    i=i+1
print("按考场总数："+str(sum))
i=1#班级标记
sum=0
while i<=Class_num[0]:
	#第i考场数据处理
    kc_value=i
    sum=sum+Class_num[i]
    kc_s=kc_value//Every_next
    kc_ys=kc_value%Every_next
    if(kc_ys==0):
        kc_s=kc_s-1
        kc_ys=Every_next
    row=int(kc_s*max_class)
    sht_bj.cell(row+1,kc_ys*5-4).value="姓名"
    sht_bj.cell(row+1,kc_ys*5-3).value="班级"
    sht_bj.cell(row+1,kc_ys*5-2).value="考场"
    sht_bj.cell(row+1,kc_ys*5-1).value="考号"
    #边框
    sht_bj.cell(row+1,kc_ys*5-4).border=border
    sht_bj.cell(row+1,kc_ys*5-3).border=border
    sht_bj.cell(row+1,kc_ys*5-2).border=border
    sht_bj.cell(row+1,kc_ys*5-1).border=border
    #print(Class_num[i])
    i=i+1
print("按班级总数："+str(sum))
#考场基本结束
dt = datetime.now()
#dt= dt.strftime( '%Y-%m-%d %H:%M:%S %f' )
dt= dt.strftime( '%Y%m%d %H_%M' )
wb.save(kao_name+dt+'.xlsx')
#wb.save("AAABBB.xlsx")  # 保存
#暂停程序，按键退出
input("任意键盘结束")





