#coding:utf-8
import openpyxl
from datetime import datetime

#基础参数设置
file_name="A0.xlsx" #汇总名单文件名
kao_name="A1" #本次考试名称
lie=6		#填充第几列
#基础参数设置OK
#设置考场和人数
Num=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,]
Stu=[41,41,41,41,41,41,41,41,41,41,41,42,42,40,40,40,40,32,40,40,40,34,34,]
print("考场总计："+str(len(Num)))
print("开始编排考号")
sum=0
t=0
while t<len(Num):
	kc=Num[t]
	rs=Stu[t]
	sum=sum+rs
	t=t+1
print("总人数统计："+str(sum))

#数组下标从0开始
wb = openpyxl.load_workbook(file_name)
ws = wb.active
#编排操作

row=2#从第几行开始
t=0

while t<len(Num):
	kc=Num[t]
	rs=Stu[t]
	cnt=1
	while cnt<=rs:
		#考场号变字符
		if(kc<10):
			kc_str='0'+str(kc)
		else:
			kc_str=str(kc)
		#座位号变字符
		if(cnt<10):
			zw='0'+str(cnt)
		else:
			zw=str(cnt)
		ws.cell(row,lie).value=kc_str
		ws.cell(row,lie+1).value=zw
		ws.cell(row,lie+2).value=kc_str+zw
		cnt=cnt+1
		row=row+1
	print("第"+str(kc)+"考场--->OK")
	t=t+1
	
#考场基本结束
dt = datetime.now()
#dt= dt.strftime( '%Y-%m-%d %H:%M:%S %f' )
dt= dt.strftime( '%Y%m%d %H_%M' )
wb.save(kao_name+'.xlsx')

print("程序运行结束")
a = input("任意键盘结束")







